from __future__ import annotations

import hashlib
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

from lore_splitter.contracts import ManifestDiagnostic
from lore_splitter.resolver import ResolvedFile
from lore_splitter.xlsx.contracts import (
    SheetExtraction,
    TableCandidate,
    WorkbookExtraction,
    WorkbookExtractionResult,
)
from lore_splitter.xlsx.merged import (
    expand_merged_values,
    extract_merged_ranges,
)
from lore_splitter.xlsx.regions import (
    build_sheet_regions,
    detect_table_candidates,
)

WARNING_HIDDEN_SHEET = "hidden_sheet"


def sha256_file(path: str | Path, *, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_workbooks(resolved_files: Iterable[ResolvedFile]) -> WorkbookExtractionResult:
    workbooks: list[WorkbookExtraction] = []
    diagnostics: list[ManifestDiagnostic] = []

    for resolved_file in resolved_files:
        try:
            workbooks.append(_extract_workbook(resolved_file))
        except Exception:  # noqa: BLE001 - record-level diagnostics must not stop batches.
            diagnostics.append(
                ManifestDiagnostic.for_source(
                    "unreadable_workbook",
                    "Could not open workbook: unreadable_workbook",
                    resolved_file.source_file,
                )
            )

    return WorkbookExtractionResult(workbooks=tuple(workbooks), diagnostics=tuple(diagnostics))


def _extract_workbook(resolved_file: ResolvedFile) -> WorkbookExtraction:
    checksum = sha256_file(resolved_file.local_path)
    workbook = load_workbook(
        resolved_file.local_path,
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    try:
        sheets = tuple(
            _extract_sheet(
                sheet,
                index=index,
                workbook_path=resolved_file.local_path,
                workbook_checksum=checksum,
            )
            for index, sheet in enumerate(workbook.worksheets, start=1)
        )
        return WorkbookExtraction(
            source_file=resolved_file.source_file,
            local_path=resolved_file.local_path,
            workbook_checksum=checksum,
            sheets=sheets,
        )
    finally:
        workbook.close()


def _extract_sheet(
    sheet,
    *,
    index: int,
    workbook_path: Path,
    workbook_checksum: str,
) -> SheetExtraction:
    hidden = getattr(sheet, "sheet_state", "visible") != "visible"
    merged_ranges = _safe_extract_merged_ranges(workbook_path, sheet.title)
    rows, merged_cells_expanded = expand_merged_values(_sheet_rows(sheet), merged_ranges)
    table_candidates = detect_table_candidates(
        rows,
        workbook_checksum=workbook_checksum,
        sheet_name=sheet.title,
        sheet_index=index,
        merged_cells_expanded=merged_cells_expanded,
    )
    if hidden:
        table_candidates = tuple(
            _with_candidate_warning(candidate, WARNING_HIDDEN_SHEET)
            for candidate in table_candidates
        )

    regions = build_sheet_regions(
        rows,
        table_candidates,
        workbook_checksum=workbook_checksum,
        sheet_name=sheet.title,
        sheet_index=index,
        merged_ranges=merged_ranges,
    )
    warnings = (WARNING_HIDDEN_SHEET,) if hidden else ()

    return SheetExtraction(
        name=sheet.title,
        index=index,
        hidden=hidden,
        max_row=sheet.max_row or 0,
        max_column=sheet.max_column or 0,
        merged_ranges=merged_ranges,
        table_candidates=table_candidates,
        regions=regions,
        warnings=warnings,
    )


def _sheet_rows(sheet) -> list[list[object]]:
    iter_rows = getattr(sheet, "iter_rows", None)
    if iter_rows is None:
        return []
    return [list(row) for row in iter_rows(values_only=True)]


def _safe_extract_merged_ranges(workbook_path: Path, sheet_name: str):
    try:
        return extract_merged_ranges(workbook_path, sheet_name)
    except Exception:  # noqa: BLE001 - merged metadata must not fail an otherwise loaded workbook.
        return ()


def _with_candidate_warning(candidate: TableCandidate, warning: str) -> TableCandidate:
    return TableCandidate(
        workbook_checksum=candidate.workbook_checksum,
        sheet_name=candidate.sheet_name,
        sheet_index=candidate.sheet_index,
        range=candidate.range,
        header_row=candidate.header_row,
        columns=candidate.columns,
        warnings=tuple(dict.fromkeys((*candidate.warnings, warning))),
    )

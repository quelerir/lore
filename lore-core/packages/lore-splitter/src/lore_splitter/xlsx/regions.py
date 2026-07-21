from __future__ import annotations

from collections.abc import Iterable, Sequence
from typing import Any

from openpyxl.utils.cell import get_column_letter

from lore_splitter.xlsx.contracts import CellRange, SheetRegion, TableCandidate

WARNING_DUPLICATE_HEADERS = "duplicate_headers"
WARNING_FALLBACK_USED = "fallback_used"
WARNING_GENERATED_HEADERS = "generated_headers"
WARNING_MERGED_CELLS_EXPANDED = "merged_cells_expanded"
WARNING_SPARSE_SHAPE = "sparse_shape"
_MIN_FALLBACK_DENSITY = 0.20
WARNING_LOW_MEANING_FRAGMENT = "low_meaning_fragment"


def detect_table_candidates(
    rows: Sequence[Sequence[Any]],
    *,
    workbook_checksum: str,
    sheet_name: str,
    sheet_index: int,
    merged_cells_expanded: bool = False,
) -> tuple[TableCandidate, ...]:
    normalized_rows = _normalize_rows(rows)
    occupied = _occupied_positions(normalized_rows)
    if not occupied:
        return ()

    row_groups = _contiguous_groups(row for row, _column in occupied)
    column_groups = _contiguous_groups(column for _row, column in occupied)
    candidates: list[TableCandidate] = []

    for row_group in row_groups:
        for column_group in column_groups:
            bounds = _trim_bounds(normalized_rows, row_group, column_group)
            if bounds is None:
                continue
            min_row, max_row, min_column, max_column = bounds
            if max_row - min_row + 1 < 2 or max_column - min_column + 1 < 2:
                continue
            header_row, columns, warnings = _infer_headers(
                normalized_rows,
                min_row=min_row,
                max_row=max_row,
                min_column=min_column,
                max_column=max_column,
            )
            candidates.append(
                TableCandidate(
                    workbook_checksum=workbook_checksum,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    range=_cell_range(min_row, max_row, min_column, max_column),
                    header_row=header_row,
                    columns=columns,
                    warnings=warnings,
                )
            )

    if candidates:
        if merged_cells_expanded:
            return tuple(
                _with_warning(candidate, WARNING_MERGED_CELLS_EXPANDED)
                for candidate in candidates
            )
        return tuple(candidates)

    return _fallback_candidate(
        normalized_rows,
        occupied,
        workbook_checksum=workbook_checksum,
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        merged_cells_expanded=merged_cells_expanded,
    )


def build_sheet_regions(
    rows: Sequence[Sequence[Any]],
    candidates: Sequence[TableCandidate],
    *,
    workbook_checksum: str,
    sheet_name: str,
    sheet_index: int,
    merged_ranges: Sequence[CellRange] = (),
) -> tuple[SheetRegion, ...]:
    """Build source-ordered table and scalar regions from cached worksheet rows."""
    normalized = _normalize_rows(rows)
    occupied = _occupied_positions(normalized)
    regions: list[SheetRegion] = []
    covered: set[tuple[int, int]] = set()
    for candidate in sorted(candidates, key=lambda item: _range_sort_key(item.range)):
        bounds = candidate.range
        candidate_rows = tuple(
            tuple(
                normalized[row - 1][column - 1]
                for column in range(bounds.min_column, bounds.max_column + 1)
            )
            for row in range(bounds.min_row, bounds.max_row + 1)
        )
        covered.update(
            (row, column)
            for row in range(bounds.min_row, bounds.max_row + 1)
            for column in range(bounds.min_column, bounds.max_column + 1)
        )
        regions.append(
            SheetRegion(
                semantic_kind="table",
                sheet_name=sheet_name,
                sheet_index=sheet_index,
                source_bounds=bounds,
                rows=candidate_rows,
                candidate=candidate,
                merged_ranges=tuple(item for item in merged_ranges if _overlaps(item, bounds)),
                warnings=candidate.warnings,
            )
        )

    scalar_rows: list[int] = []
    for row_index, row in enumerate(normalized, start=1):
        if any(
            (row_index, column) in occupied and (row_index, column) not in covered
            for column in range(1, len(row) + 1)
        ):
            scalar_rows.append(row_index)
    for group in _contiguous_groups(scalar_rows):
        cells = [
            (row, column)
            for row in group
            for column in range(1, len(normalized[row - 1]) + 1)
            if (row, column) in occupied and (row, column) not in covered
        ]
        if not cells:
            continue
        min_row, max_row = group.start, group.stop - 1
        min_column = min(column for _row, column in cells)
        max_column = max(column for _row, column in cells)
        bounds = _cell_range(min_row, max_row, min_column, max_column)
        values = [str(normalized[row - 1][column - 1]).strip() for row, column in cells]
        text = "\n".join(value for value in values if value)
        semantic_kind = (
            "scalar"
            if len(cells) > 1 or (min_row == 1 and len(text) >= 3) or len(text) >= 12
            else "skipped"
        )
        warnings = () if semantic_kind == "scalar" else (WARNING_LOW_MEANING_FRAGMENT,)
        regions.append(
            SheetRegion(
                semantic_kind=semantic_kind,
                sheet_name=sheet_name,
                sheet_index=sheet_index,
                source_bounds=bounds,
                text=text,
                rows=tuple((normalized[row - 1][column - 1],) for row, column in cells),
                merged_ranges=tuple(item for item in merged_ranges if _overlaps(item, bounds)),
                warnings=warnings,
            )
        )
    return tuple(sorted(regions, key=lambda region: _range_sort_key(region.source_bounds)))


def _range_sort_key(bounds: CellRange) -> tuple[int, int, int, int]:
    return bounds.min_row, bounds.min_column, bounds.max_row, bounds.max_column


def _overlaps(left: CellRange, right: CellRange) -> bool:
    return not (
        left.max_row < right.min_row
        or right.max_row < left.min_row
        or left.max_column < right.min_column
        or right.max_column < left.min_column
    )


def _normalize_rows(rows: Sequence[Sequence[Any]]) -> list[list[Any]]:
    width = max((len(row) for row in rows), default=0)
    return [list(row) + [None] * (width - len(row)) for row in rows]


def _occupied_positions(rows: Sequence[Sequence[Any]]) -> set[tuple[int, int]]:
    occupied: set[tuple[int, int]] = set()
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            if not _is_blank(value):
                occupied.add((row_index, column_index))
    return occupied


def _contiguous_groups(values: Iterable[int]) -> list[range]:
    groups: list[range] = []
    sorted_values = sorted(set(values))
    if not sorted_values:
        return groups

    start = previous = sorted_values[0]
    for value in sorted_values[1:]:
        if value == previous + 1:
            previous = value
            continue
        groups.append(range(start, previous + 1))
        start = previous = value
    groups.append(range(start, previous + 1))
    return groups


def _trim_bounds(
    rows: Sequence[Sequence[Any]],
    row_group: range,
    column_group: range,
) -> tuple[int, int, int, int] | None:
    min_row = row_group.start
    max_row = row_group.stop - 1
    min_column = column_group.start
    max_column = column_group.stop - 1

    while min_row <= max_row and _occupied_count_in_row(rows, min_row, min_column, max_column) < 2:
        min_row += 1
    while max_row >= min_row and _occupied_count_in_row(rows, max_row, min_column, max_column) < 2:
        max_row -= 1
    while (
        min_column <= max_column
        and _occupied_count_in_column(rows, min_column, min_row, max_row) < 2
    ):
        min_column += 1
    while (
        max_column >= min_column
        and _occupied_count_in_column(rows, max_column, min_row, max_row) < 2
    ):
        max_column -= 1

    if min_row > max_row or min_column > max_column:
        return None
    return min_row, max_row, min_column, max_column


def _occupied_count_in_row(
    rows: Sequence[Sequence[Any]], row_index: int, min_column: int, max_column: int
) -> int:
    row = rows[row_index - 1]
    return sum(1 for column in range(min_column, max_column + 1) if not _is_blank(row[column - 1]))


def _occupied_count_in_column(
    rows: Sequence[Sequence[Any]], column_index: int, min_row: int, max_row: int
) -> int:
    return sum(
        1
        for row_index in range(min_row, max_row + 1)
        if not _is_blank(rows[row_index - 1][column_index - 1])
    )


def _infer_headers(
    rows: Sequence[Sequence[Any]],
    *,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> tuple[int, tuple[str, ...], tuple[str, ...]]:
    header_row = min_row
    for row_index in range(min_row, min(max_row, min_row + 2) + 1):
        values = [
            rows[row_index - 1][column_index - 1]
            for column_index in range(min_column, max_column + 1)
        ]
        if _textual_header_score(values) >= max(1, len(values) // 2):
            header_row = row_index
            break

    raw_labels = [
        rows[header_row - 1][column_index - 1]
        for column_index in range(min_column, max_column + 1)
    ]
    columns, warnings = _stable_column_labels(raw_labels)
    return header_row, columns, warnings


def _fallback_candidate(
    rows: Sequence[Sequence[Any]],
    occupied: set[tuple[int, int]],
    *,
    workbook_checksum: str,
    sheet_name: str,
    sheet_index: int,
    merged_cells_expanded: bool,
) -> tuple[TableCandidate, ...]:
    if not occupied:
        return ()
    min_row = min(row for row, _column in occupied)
    max_row = max(row for row, _column in occupied)
    min_column = min(column for _row, column in occupied)
    max_column = max(column for _row, column in occupied)
    if _occupied_count_in_row(rows, min_row, min_column, max_column) < 2:
        return ()
    if len(occupied) / (max_row * max_column) < _MIN_FALLBACK_DENSITY:
        return ()

    header_row, columns, header_warnings = _infer_headers(
        rows,
        min_row=min_row,
        max_row=max_row,
        min_column=min_column,
        max_column=max_column,
    )
    warnings = [WARNING_FALLBACK_USED, *header_warnings]
    if _is_sparse(occupied, min_row, max_row, min_column, max_column):
        warnings.append(WARNING_SPARSE_SHAPE)
    if merged_cells_expanded:
        warnings.append(WARNING_MERGED_CELLS_EXPANDED)

    return (
        TableCandidate(
            workbook_checksum=workbook_checksum,
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            range=_cell_range(min_row, max_row, min_column, max_column),
            header_row=header_row,
            columns=columns,
            warnings=tuple(dict.fromkeys(warnings)),
        ),
    )


def _is_sparse(
    occupied: set[tuple[int, int]],
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> bool:
    area = (max_row - min_row + 1) * (max_column - min_column + 1)
    if area == 0:
        return False
    return len(occupied) / area < 0.5


def _with_warning(candidate: TableCandidate, warning: str) -> TableCandidate:
    return TableCandidate(
        workbook_checksum=candidate.workbook_checksum,
        sheet_name=candidate.sheet_name,
        sheet_index=candidate.sheet_index,
        range=candidate.range,
        header_row=candidate.header_row,
        columns=candidate.columns,
        warnings=tuple(dict.fromkeys((*candidate.warnings, warning))),
    )


def _stable_column_labels(values: Sequence[Any]) -> tuple[tuple[str, ...], tuple[str, ...]]:
    warnings: list[str] = []
    labels: list[str] = []
    seen: dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        if _is_blank(value):
            base_label = f"Column_{index}"
            if WARNING_GENERATED_HEADERS not in warnings:
                warnings.append(WARNING_GENERATED_HEADERS)
        else:
            base_label = str(value).strip()

        occurrence = seen.get(base_label, 0) + 1
        seen[base_label] = occurrence
        if occurrence > 1:
            label = f"{base_label}_{occurrence}"
            if WARNING_DUPLICATE_HEADERS not in warnings:
                warnings.append(WARNING_DUPLICATE_HEADERS)
        else:
            label = base_label
        labels.append(label)

    return tuple(labels), tuple(warnings)


def _textual_header_score(values: Sequence[Any]) -> int:
    return sum(1 for value in values if isinstance(value, str) and value.strip())


def _cell_range(min_row: int, max_row: int, min_column: int, max_column: int) -> CellRange:
    a1_range = f"{get_column_letter(min_column)}{min_row}:{get_column_letter(max_column)}{max_row}"
    return CellRange(
        min_row=min_row,
        max_row=max_row,
        min_column=min_column,
        max_column=max_column,
        a1_range=a1_range,
    )


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")

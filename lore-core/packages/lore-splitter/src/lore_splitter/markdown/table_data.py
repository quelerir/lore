from __future__ import annotations

from collections.abc import Iterable

from openpyxl import load_workbook

from lore_splitter.contracts import ManifestDiagnostic
from lore_splitter.markdown.contracts import (
    TableData,
    TableDataExtractionResult,
    XlsxTableLocation,
)
from lore_splitter.xlsx.contracts import (
    SheetExtraction,
    TableCandidate,
    WorkbookExtraction,
)


def extract_table_data(workbooks: Iterable[WorkbookExtraction]) -> TableDataExtractionResult:
    tables: list[TableData] = []
    diagnostics: list[ManifestDiagnostic] = []

    for workbook in workbooks:
        try:
            tables.extend(_extract_workbook_tables(workbook))
        except Exception as exc:  # noqa: BLE001 - one workbook must not stop the batch.
            diagnostics.append(
                ManifestDiagnostic.for_source(
                    "unreadable_table_values",
                    f"Could not read workbook table values: {exc}",
                    workbook.source_file,
                )
            )

    return TableDataExtractionResult(tables=tuple(tables), diagnostics=tuple(diagnostics))


def _extract_workbook_tables(workbook: WorkbookExtraction) -> list[TableData]:
    loaded = load_workbook(
        workbook.local_path,
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    try:
        tables: list[TableData] = []
        table_index = 1
        for sheet in workbook.sheets:
            if not sheet.table_candidates:
                continue
            worksheet = _worksheet_for_sheet(loaded, sheet)
            for candidate in sheet.table_candidates:
                tables.append(_extract_candidate_table(workbook, candidate, worksheet, table_index))
                table_index += 1
        return tables
    finally:
        loaded.close()


def _worksheet_for_sheet(loaded_workbook, sheet: SheetExtraction):
    try:
        worksheet = loaded_workbook.worksheets[sheet.index - 1]
    except IndexError:
        worksheet = loaded_workbook[sheet.name]

    if worksheet.title != sheet.name:
        worksheet = loaded_workbook[sheet.name]
    return worksheet


def _extract_candidate_table(
    workbook: WorkbookExtraction,
    candidate: TableCandidate,
    worksheet,
    table_index: int,
) -> TableData:
    cell_range = candidate.range
    rows = tuple(
        tuple(row)
        for row in worksheet.iter_rows(
            min_row=cell_range.min_row,
            max_row=cell_range.max_row,
            min_col=cell_range.min_column,
            max_col=cell_range.max_column,
            values_only=True,
        )
    )
    return TableData(
        source_file=workbook.source_file,
        local_path=workbook.local_path,
        source_kind="workbook",
        source_checksum=workbook.workbook_checksum,
        table_index=table_index,
        columns=candidate.columns,
        rows=rows,
        xlsx=XlsxTableLocation(
            workbook_checksum=workbook.workbook_checksum,
            sheet_name=candidate.sheet_name,
            sheet_index=candidate.sheet_index,
            range=candidate.range,
            header_row=candidate.header_row,
        ),
        warnings=candidate.warnings,
    )

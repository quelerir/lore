import hashlib
import json

from lore_splitter.manifest import load_manifest
from lore_splitter.resolver import ResolvedFile, resolve_manifest
from lore_splitter.xlsx import extract_workbooks, sha256_file
from lore_splitter.xlsx import workbook as workbook_module

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"


def test_sha256_file_hashes_exact_local_bytes(tmp_path) -> None:
    local_file = tmp_path / "bytes.xlsx"
    content = b"exact workbook bytes"
    local_file.write_bytes(content)

    assert sha256_file(local_file) == hashlib.sha256(content).hexdigest()


def test_extract_workbooks_parses_manifest_resolved_xlsx_metadata(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "report.xlsx"
    _write_workbook(workbook_path, suffix=".xlsx", hidden_second_sheet=True)
    manifest_path = _write_manifest(
        tmp_path,
        [
            {
                "source_id": "google-drive",
                "stream": "regulations",
                "file_id": "xlsx-1",
                "source_path": "Finance/report.xlsx",
                "object_path": "/staging/files/report.xlsx",
                "mime_type": XLSX_MIME,
                "size_bytes": workbook_path.stat().st_size,
            }
        ],
    )

    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)

    assert result.diagnostics == ()
    assert len(result.workbooks) == 1
    workbook = result.workbooks[0]
    assert workbook.file_id == "xlsx-1"
    assert workbook.workbook_checksum == sha256_file(workbook_path)
    assert [sheet.name for sheet in workbook.sheets] == ["Summary", "Lookup"]
    assert workbook.sheets[0].index == 1
    assert workbook.sheets[0].hidden is False
    assert workbook.sheets[0].max_row == 3
    assert workbook.sheets[0].max_column == 3
    assert [cell_range.a1_range for cell_range in workbook.sheets[0].merged_ranges] == []
    assert workbook.sheets[1].index == 2
    assert workbook.sheets[1].hidden is True


def test_extract_workbooks_parses_manifest_resolved_xlsm_with_same_path(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "macro.xlsm"
    _write_workbook(workbook_path, suffix=".xlsm")
    manifest_path = _write_manifest(
        tmp_path,
        [
            {
                "source_id": "google-drive",
                "stream": "regulations",
                "file_id": "xlsm-1",
                "source_path": "Finance/macro.xlsm",
                "object_path": "/staging/files/macro.xlsm",
                "mime_type": XLSM_MIME,
                "size_bytes": workbook_path.stat().st_size,
            }
        ],
    )

    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)

    assert result.diagnostics == ()
    assert len(result.workbooks) == 1
    assert result.workbooks[0].file_id == "xlsm-1"
    assert result.workbooks[0].sheets[0].name == "Summary"


def test_extract_workbooks_returns_diagnostic_for_unreadable_workbook(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    valid_path = input_root / "staging" / "files" / "valid.xlsx"
    corrupt_path = input_root / "staging" / "files" / "corrupt.xlsx"
    _write_workbook(valid_path, suffix=".xlsx")
    corrupt_path.parent.mkdir(parents=True, exist_ok=True)
    corrupt_path.write_bytes(b"not an xlsx zip")
    manifest_path = _write_manifest(
        tmp_path,
        [
            {
                "source_id": "google-drive",
                "stream": "regulations",
                "file_id": "valid-1",
                "source_path": "valid.xlsx",
                "object_path": "/staging/files/valid.xlsx",
                "mime_type": XLSX_MIME,
                "size_bytes": valid_path.stat().st_size,
            },
            {
                "source_id": "google-drive",
                "stream": "regulations",
                "file_id": "corrupt-1",
                "source_path": "corrupt.xlsx",
                "object_path": "/staging/files/corrupt.xlsx",
                "mime_type": XLSX_MIME,
                "size_bytes": corrupt_path.stat().st_size,
            },
        ],
    )

    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)

    assert [workbook.file_id for workbook in result.workbooks] == ["valid-1"]
    assert len(result.diagnostics) == 1
    diagnostic = result.diagnostics[0]
    assert diagnostic.reason == "unreadable_workbook"
    assert diagnostic.file_id == "corrupt-1"
    assert "Could not open workbook" in diagnostic.message


def test_extract_workbooks_uses_read_only_data_only_and_closes(monkeypatch, tmp_path) -> None:
    source_path = tmp_path / "input.xlsx"
    source_path.write_bytes(b"fake bytes")
    source_file = _source_file("xlsx-1", "/staging/files/input.xlsx")
    calls = {}

    class FakeSheet:
        title = "Summary"
        sheet_state = "visible"
        max_row = 4
        max_column = 2

    class FakeWorkbook:
        worksheets = [FakeSheet()]

        def close(self) -> None:
            calls["closed"] = True

    def fake_load_workbook(path, *, read_only, data_only, keep_vba):
        calls["path"] = path
        calls["read_only"] = read_only
        calls["data_only"] = data_only
        calls["keep_vba"] = keep_vba
        return FakeWorkbook()

    monkeypatch.setattr(workbook_module, "load_workbook", fake_load_workbook)

    result = extract_workbooks((ResolvedFile(source_file=source_file, local_path=source_path),))

    assert result.diagnostics == ()
    assert result.workbooks[0].sheets[0].name == "Summary"
    assert calls == {
        "path": source_path,
        "read_only": True,
        "data_only": True,
        "keep_vba": False,
        "closed": True,
    }


def test_extract_workbooks_attaches_table_candidates_to_visible_sheet(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "report.xlsx"
    _write_workbook(workbook_path, suffix=".xlsx")
    workbook = _extract_single_workbook(tmp_path, input_root, workbook_path)

    candidates = workbook.sheets[0].table_candidates

    assert len(candidates) == 1
    candidate = candidates[0]
    assert candidate.workbook_checksum == workbook.workbook_checksum
    assert candidate.sheet_name == "Summary"
    assert candidate.sheet_index == 1
    assert candidate.range.a1_range == "A1:C3"
    assert candidate.header_row == 1
    assert candidate.columns == ("Name", "Amount", "Date")
    assert candidate.warnings == ()


def test_extract_workbooks_keeps_empty_sheets_without_candidates_and_marks_hidden_candidates(
    tmp_path,
) -> None:
    from openpyxl import Workbook

    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "hidden.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    empty = workbook.active
    empty.title = "Empty"
    hidden = workbook.create_sheet("HiddenData")
    hidden.sheet_state = "hidden"
    hidden.append(["Code", "Label"])
    hidden.append(["A", "Active"])
    workbook.save(workbook_path)
    workbook.close()

    extraction = _extract_single_workbook(tmp_path, input_root, workbook_path)

    assert extraction.sheets[0].name == "Empty"
    assert extraction.sheets[0].table_candidates == ()
    assert extraction.sheets[1].hidden is True
    assert len(extraction.sheets[1].table_candidates) == 1
    assert "hidden_sheet" in extraction.sheets[1].table_candidates[0].warnings


def test_extract_workbooks_preserves_merged_ranges_and_expands_candidates(tmp_path) -> None:
    from openpyxl import Workbook

    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "merged.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Merged"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Region"
    sheet["C1"] = "Amount"
    sheet.append(["North", "North", 10])
    sheet.append(["South", "South", 20])
    workbook.save(workbook_path)
    workbook.close()

    extraction = _extract_single_workbook(tmp_path, input_root, workbook_path)

    sheet = extraction.sheets[0]
    assert [cell_range.a1_range for cell_range in sheet.merged_ranges] == ["A1:B1"]
    assert len(sheet.table_candidates) == 1
    assert sheet.table_candidates[0].range.a1_range == "A1:C3"
    assert "merged_cells_expanded" in sheet.table_candidates[0].warnings


def test_extract_workbooks_uses_fallback_for_ambiguous_data_sheet(tmp_path) -> None:
    from openpyxl import Workbook

    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "fallback.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet.append(["Only one populated row", "with context", "but no body"])
    sheet.append([None, None, None])
    sheet.append([None, "Sparse note", None])
    workbook.save(workbook_path)
    workbook.close()

    extraction = _extract_single_workbook(tmp_path, input_root, workbook_path)

    candidates = extraction.sheets[0].table_candidates
    assert len(candidates) == 1
    assert candidates[0].range.a1_range == "A1:C3"
    assert "fallback_used" in candidates[0].warnings
    assert "sparse_shape" in candidates[0].warnings


def _extract_single_workbook(tmp_path, input_root, workbook_path):
    manifest_path = _write_manifest(
        tmp_path,
        [
            {
                "source_id": "google-drive",
                "stream": "regulations",
                "file_id": workbook_path.stem,
                "source_path": workbook_path.name,
                "object_path": f"/staging/files/{workbook_path.name}",
                "mime_type": XLSX_MIME,
                "size_bytes": workbook_path.stat().st_size,
            }
        ],
    )
    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)
    assert result.diagnostics == ()
    return result.workbooks[0]


def _source_file(file_id: str, object_path: str):
    from lore_splitter.contracts import SourceFile

    return SourceFile(
        source_id="google-drive",
        stream="regulations",
        file_id=file_id,
        source_path=object_path.rsplit("/", 1)[-1],
        object_path=object_path,
        mime_type=XLSX_MIME,
        size_bytes=100,
    )


def _write_manifest(tmp_path, records) -> object:
    manifest_path = tmp_path / "manifest.jsonl"
    manifest_path.write_text("\n".join(json.dumps(record) for record in records), encoding="utf-8")
    return manifest_path


def _write_workbook(path, *, suffix: str, hidden_second_sheet: bool = False) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first = workbook.active
    first.title = "Summary"
    first.append(["Name", "Amount", "Date"])
    first.append(["Alpha", 10, "2026-07-01"])
    first.append(["Beta", 20, "2026-07-02"])
    second = workbook.create_sheet("Lookup")
    second.append(["Code", "Label"])
    second.append(["A", "Alpha"])
    if hidden_second_sheet:
        second.sheet_state = "hidden"
    workbook.save(path.with_suffix(suffix))
    workbook.close()

from __future__ import annotations

import json

from lore_splitter.manifest import load_manifest
from lore_splitter.resolver import resolve_manifest
from lore_splitter.xlsx import extract_workbooks
from openpyxl import Workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_generated_edge_case_workbook_reports_exact_ranges_and_warnings(tmp_path) -> None:
    input_root = tmp_path / "fixtures"
    workbook_path = input_root / "staging" / "files" / "edge-cases.xlsx"
    _write_edge_case_workbook(workbook_path)

    workbook = _extract_single_workbook(tmp_path, input_root, workbook_path)
    sheets = {sheet.name: sheet for sheet in workbook.sheets}

    assert _candidate_ranges(sheets["SingleTable"]) == ["A1:C3"]
    assert _candidate_ranges(sheets["TitlePlusTable"]) == ["A3:C5"]
    assert _candidate_ranges(sheets["MultipleTables"]) == ["A1:B3", "D1:E3"]

    merged = sheets["MergedCells"]
    assert [cell_range.a1_range for cell_range in merged.merged_ranges] == ["A1:B1"]
    assert _candidate_ranges(merged) == ["A1:C3"]
    assert "merged_cells_expanded" in merged.table_candidates[0].warnings

    duplicate_headers = sheets["DuplicateHeaders"].table_candidates[0]
    assert duplicate_headers.columns == ("Name", "Name_2", "Column_3")
    assert duplicate_headers.warnings == ("duplicate_headers", "generated_headers")

    formulas = sheets["Formulas"].table_candidates[0]
    assert formulas.range.a1_range == "A1:B2"
    assert formulas.columns == ("Label", "Total")

    hidden = sheets["HiddenLookup"]
    assert hidden.hidden is True
    assert hidden.table_candidates[0].range.a1_range == "A1:B2"
    assert "hidden_sheet" in hidden.table_candidates[0].warnings


def test_sparse_sheet_uses_whole_sheet_fallback_with_exact_coordinates(tmp_path) -> None:
    input_root = tmp_path / "fixtures"
    workbook_path = input_root / "staging" / "files" / "sparse.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SparseFallback"
    sheet["A1"] = "Sparse workbook"
    sheet["D1"] = "context"
    sheet["B3"] = "important note"
    sheet["D4"] = "owner"
    workbook.save(workbook_path)
    workbook.close()

    extraction = _extract_single_workbook(tmp_path, input_root, workbook_path)

    candidates = extraction.sheets[0].table_candidates
    assert len(candidates) == 1
    assert candidates[0].range.a1_range == "A1:D4"
    assert "fallback_used" in candidates[0].warnings
    assert "sparse_shape" in candidates[0].warnings


def test_large_sparse_and_dense_workbooks_parse_through_read_only_path(tmp_path) -> None:
    input_root = tmp_path / "fixtures"
    sparse_path = input_root / "staging" / "files" / "large-sparse.xlsx"
    dense_path = input_root / "staging" / "files" / "large-dense.xlsx"
    _write_large_sparse_workbook(sparse_path)
    _write_large_dense_workbook(dense_path)
    manifest_path = _write_manifest(
        tmp_path,
        [
            _manifest_record("large-sparse", sparse_path),
            _manifest_record("large-dense", dense_path),
        ],
    )

    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)

    assert result.diagnostics == ()
    assert [workbook.file_id for workbook in result.workbooks] == [
        "large-sparse",
        "large-dense",
    ]
    assert result.workbooks[0].sheets[0].max_row == 300
    assert result.workbooks[0].sheets[0].max_column == 25
    assert result.workbooks[1].sheets[0].table_candidates[0].range.a1_range == "A1:F80"


def _extract_single_workbook(tmp_path, input_root, workbook_path):
    manifest_path = _write_manifest(tmp_path, [_manifest_record(workbook_path.stem, workbook_path)])
    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)
    assert result.diagnostics == ()
    assert len(result.workbooks) == 1
    return result.workbooks[0]


def _write_edge_case_workbook(path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    single = workbook.active
    single.title = "SingleTable"
    single.append(["Name", "Amount", "Owner"])
    single.append(["North", 10, "Ada"])
    single.append(["South", 20, "Grace"])

    title = workbook.create_sheet("TitlePlusTable")
    title.append(["Quarterly sales"])
    title.append([])
    title.append(["Region", "Amount", "Owner"])
    title.append(["North", 10, "Ada"])
    title.append(["South", 20, "Grace"])

    multiple = workbook.create_sheet("MultipleTables")
    multiple.append(["Region", "Amount", None, "Code", "Label"])
    multiple.append(["North", 10, None, "A", "Active"])
    multiple.append(["South", 20, None, "I", "Inactive"])

    merged = workbook.create_sheet("MergedCells")
    merged.merge_cells("A1:B1")
    merged["A1"] = "Region"
    merged["C1"] = "Amount"
    merged.append(["North", "North", 10])
    merged.append(["South", "South", 20])

    duplicate = workbook.create_sheet("DuplicateHeaders")
    duplicate.append(["Name", "Name", None])
    duplicate.append(["Alpha", "A", 10])
    duplicate.append(["Beta", "B", 20])

    formulas = workbook.create_sheet("Formulas")
    formulas.append(["Label", "Total"])
    formulas.append(["North", 10])
    formulas.append(["South", "=B2*2"])

    hidden = workbook.create_sheet("HiddenLookup")
    hidden.sheet_state = "hidden"
    hidden.append(["Code", "Label"])
    hidden.append(["A", "Active"])

    workbook.save(path)
    workbook.close()


def _write_large_sparse_workbook(path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LargeSparse"
    sheet["A1"] = "Anchor"
    sheet["Y1"] = "Far"
    sheet["B300"] = "Tail"
    workbook.save(path)
    workbook.close()


def _write_large_dense_workbook(path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("LargeDense")
    sheet.append(["Col1", "Col2", "Col3", "Col4", "Col5", "Col6"])
    for row_index in range(1, 80):
        sheet.append([f"row-{row_index}", row_index, row_index + 1, row_index + 2, "x", "y"])
    workbook.save(path)
    workbook.close()


def _manifest_record(file_id: str, workbook_path) -> dict[str, object]:
    return {
        "source_id": "google-drive",
        "stream": "regulations",
        "file_id": file_id,
        "source_path": workbook_path.name,
        "object_path": f"/staging/files/{workbook_path.name}",
        "mime_type": XLSX_MIME,
        "size_bytes": workbook_path.stat().st_size,
    }


def _write_manifest(tmp_path, records) -> object:
    manifest_path = tmp_path / "xlsx_manifest.jsonl"
    manifest_path.write_text("\n".join(json.dumps(record) for record in records), encoding="utf-8")
    return manifest_path


def _candidate_ranges(sheet) -> list[str]:
    return [candidate.range.a1_range for candidate in sheet.table_candidates]

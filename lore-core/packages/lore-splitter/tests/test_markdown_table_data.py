from __future__ import annotations

import json

from lore_splitter.manifest import load_manifest
from lore_splitter.markdown import extract_table_data
from lore_splitter.markdown import table_data as table_data_module
from lore_splitter.resolver import resolve_manifest
from lore_splitter.xlsx import extract_workbooks
from openpyxl import Workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_extract_table_data_slices_title_plus_table_candidate_values(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "title.xlsx"
    _write_title_plus_table_workbook(workbook_path)
    workbook = _extract_single_workbook(tmp_path, input_root, workbook_path)

    result = extract_table_data((workbook,))

    assert result.diagnostics == ()
    assert len(result.tables) == 1
    table = result.tables[0]
    assert table.sheet_name == "TitlePlusTable"
    assert table.range.a1_range == "A3:C5"
    assert table.header_row == 3
    assert table.columns == ("Region", "Amount", "Owner")
    assert table.rows == (
        ("Region", "Amount", "Owner"),
        ("North", 10, "Ada"),
        ("South", 20, "Grace"),
    )
    assert table.workbook_checksum == workbook.workbook_checksum
    assert table.source_file.identity_dict() == workbook.source_file.identity_dict()


def test_extract_table_data_preserves_sheet_and_candidate_order(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "multiple.xlsx"
    _write_multiple_table_workbook(workbook_path)
    workbook = _extract_single_workbook(tmp_path, input_root, workbook_path)

    result = extract_table_data((workbook,))

    assert result.diagnostics == ()
    assert [(table.sheet_name, table.range.a1_range) for table in result.tables] == [
        ("First", "A1:B2"),
        ("Second", "A1:B2"),
        ("Second", "D1:E2"),
    ]
    assert result.tables[1].rows == (("Region", "Amount"), ("North", 10))
    assert result.tables[2].rows == (("Code", "Label"), ("A", "Active"))


def test_extract_table_data_retains_candidate_warnings(tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "warnings.xlsx"
    _write_warning_workbook(workbook_path)
    workbook = _extract_single_workbook(tmp_path, input_root, workbook_path)

    result = extract_table_data((workbook,))

    warning_map = {table.sheet_name: table.warnings for table in result.tables}
    assert "merged_cells_expanded" in warning_map["Merged"]
    assert warning_map["Duplicate"] == ("duplicate_headers", "generated_headers")
    assert "hidden_sheet" in warning_map["Hidden"]


def test_extract_table_data_returns_diagnostic_when_workbook_values_cannot_be_read(
    tmp_path,
) -> None:
    input_root = tmp_path / "materialized"
    valid_path = input_root / "staging" / "files" / "valid.xlsx"
    deleted_path = input_root / "staging" / "files" / "deleted.xlsx"
    _write_title_plus_table_workbook(valid_path)
    _write_title_plus_table_workbook(deleted_path)
    workbooks = _extract_workbooks(tmp_path, input_root, (valid_path, deleted_path))
    deleted_path.unlink()

    result = extract_table_data(workbooks)

    assert [table.file_id for table in result.tables] == ["valid"]
    assert len(result.diagnostics) == 1
    diagnostic = result.diagnostics[0]
    assert diagnostic.reason == "unreadable_table_values"
    assert diagnostic.file_id == "deleted"
    assert "Could not read workbook table values" in diagnostic.message


def test_extract_table_data_uses_read_only_data_only_and_closes(monkeypatch, tmp_path) -> None:
    input_root = tmp_path / "materialized"
    workbook_path = input_root / "staging" / "files" / "flags.xlsx"
    _write_title_plus_table_workbook(workbook_path)
    workbook = _extract_single_workbook(tmp_path, input_root, workbook_path)
    calls = {}

    class FakeSheet:
        title = "TitlePlusTable"

        def iter_rows(self, *, min_row, max_row, min_col, max_col, values_only):
            calls["bounds"] = (min_row, max_row, min_col, max_col, values_only)
            return iter(
                [
                    ("Region", "Amount", "Owner"),
                    ("North", 10, "Ada"),
                    ("South", 20, "Grace"),
                ]
            )

    class FakeWorkbook:
        worksheets = [FakeSheet()]

        def close(self) -> None:
            calls["closed"] = True

    def fake_load_workbook(path, *, read_only, data_only, keep_vba):
        calls["path"] = path
        calls["read_only"] = read_only
        calls["data_only"] = data_only
        calls["keep_vba"] = keep_vba
        return FakeWorkbook()

    monkeypatch.setattr(table_data_module, "load_workbook", fake_load_workbook)

    result = extract_table_data((workbook,))

    assert result.diagnostics == ()
    assert result.tables[0].rows[1] == ("North", 10, "Ada")
    assert calls == {
        "path": workbook_path,
        "read_only": True,
        "data_only": True,
        "keep_vba": False,
        "bounds": (3, 5, 1, 3, True),
        "closed": True,
    }


def _extract_single_workbook(tmp_path, input_root, workbook_path):
    result = _extract_workbooks(tmp_path, input_root, (workbook_path,))
    assert len(result) == 1
    return result[0]


def _extract_workbooks(tmp_path, input_root, workbook_paths):
    manifest_path = _write_manifest(
        tmp_path,
        [_manifest_record(workbook_path.stem, workbook_path) for workbook_path in workbook_paths],
    )
    resolved = resolve_manifest(load_manifest(manifest_path), input_root)
    result = extract_workbooks(resolved.processable)
    assert result.diagnostics == ()
    return result.workbooks


def _write_title_plus_table_workbook(path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TitlePlusTable"
    sheet.append(["Quarterly sales"])
    sheet.append([])
    sheet.append(["Region", "Amount", "Owner"])
    sheet.append(["North", 10, "Ada"])
    sheet.append(["South", 20, "Grace"])
    workbook.save(path)
    workbook.close()


def _write_multiple_table_workbook(path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Name", "Score"])
    first.append(["Alpha", 1])
    second = workbook.create_sheet("Second")
    second.append(["Region", "Amount", None, "Code", "Label"])
    second.append(["North", 10, None, "A", "Active"])
    workbook.save(path)
    workbook.close()


def _write_warning_workbook(path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    merged = workbook.active
    merged.title = "Merged"
    merged.merge_cells("A1:B1")
    merged["A1"] = "Region"
    merged["C1"] = "Amount"
    merged.append(["North", "North", 10])
    duplicate = workbook.create_sheet("Duplicate")
    duplicate.append(["Name", "Name", None])
    duplicate.append(["Alpha", "A", 1])
    duplicate.append(["Beta", "B", 2])
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden.append(["Code", "Label"])
    hidden.append(["A", "Active"])
    workbook.save(path)
    workbook.close()


def _manifest_record(file_id: str, workbook_path) -> dict[str, object]:
    return {
        "source_id": "google-drive",
        "stream": "regulations",
        "file_id": file_id,
        "source_path": workbook_path.name,
        "object_path": f"/staging/files/{workbook_path.name}",
        "mime_type": XLSX_MIME,
        "size_bytes": workbook_path.stat().st_size,
    }


def _write_manifest(tmp_path, records) -> object:
    manifest_path = tmp_path / "xlsx_manifest.jsonl"
    manifest_path.write_text("\n".join(json.dumps(record) for record in records), encoding="utf-8")
    return manifest_path
